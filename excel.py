from openpyxl import Workbook
from datetime import datetime
import calendar

# Diccionario para mapear nombres de meses en inglés a español
meses_espanol = {
    "January": "enero",
    "February": "febrero",
    "March": "marzo",
    "April": "abril",
    "May": "mayo",
    "June": "junio",
    "July": "julio",
    "August": "agosto",
    "September": "septiembre",
    "October": "octubre",
    "November": "noviembre",
    "December": "diciembre"
}

# Función para agregar una nueva fecha y espacio para asistencia
def agregar_fecha_y_asistencia(ws, fecha, estudiantes):
    # Obtener la última columna con datos
    ultima_columna = ws.max_column

    # Fusionar dos celdas para la nueva fecha
    ws.merge_cells(start_row=1, start_column=ultima_columna + 1, end_row=1, end_column=ultima_columna + 2)

    # Agregar la nueva fecha en la celda fusionada
    ws.cell(row=1, column=ultima_columna + 1, value=fecha)

    # Agregar espacio para la asistencia de cada estudiante
    for i, estudiante in enumerate(estudiantes, start=3):
        ws.cell(row=i, column=ultima_columna + 1, value="")

# Crear un nuevo libro de trabajo
wb = Workbook()

# Obtener la hoja activa
ws = wb.active

# Obtener la fecha actual
fecha_actual = datetime.now().strftime("%Y-%m-%d")



# Insertar la fecha actual en la parte superior del archivo
ws.merge_cells('A1:B1')
ws['A1'] = f'Fecha: {fecha_actual}'

# Agregar nombres de estudiantes
estudiantes = ["Juan", "María", "Pedro", "Ana", "Luis"]



# Agregar nombres de estudiantes a las filas
for i, estudiante in enumerate(estudiantes, start=3):
    ws.cell(row=i, column=1, value=estudiante)

# Obtener el nombre del mes actual en español
mes_actual = datetime.now().strftime("%B")
nombre_mes_actual = meses_espanol.get(mes_actual)

# Agregar una fecha y espacio para la asistencia
agregar_fecha_y_asistencia(ws, fecha_actual, estudiantes)

# Guardar el archivo con el nombre del mes actual en español
nombre_archivo = f"asistencia_{nombre_mes_actual}.xlsx"

# Guardar el archivo
wb.save(nombre_archivo)
