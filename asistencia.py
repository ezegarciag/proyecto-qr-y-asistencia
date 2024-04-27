import pandas as pd
from openpyxl import load_workbook
from datetime import datetime





def marca_asistencia(ws, nombre):
    # Obtener la fecha y hora actual
    fecha_actual = datetime.now()

    # Convertir la fecha y hora a una cadena de texto en un formato específico
    fecha_string = fecha_actual.strftime("%d/%m/%Y")
    
    # Buscar la columna correspondiente a la fecha actual
    columna_fecha = None
    for col in range(2, ws.max_column + 1):
        # Obtener el valor de la celda
        fecha_celda = ws.cell(row=1, column=col).value
        # Verificar si el valor de la celda es una cadena de texto
        if isinstance(fecha_celda, str):
            # Comparar la fecha de la hoja con la fecha actual
            if fecha_celda == fecha_string:
                columna_fecha = col
                break
        elif isinstance(fecha_celda, datetime):
            # Convertir la fecha de la celda a cadena de texto y comparar con la fecha actual
            fecha_hoja = fecha_celda.strftime('%d/%m/%Y')
            if fecha_hoja == fecha_string:
                columna_fecha = col
                break

    # Si no se encontró la columna para la fecha actual, salir de la función
    if columna_fecha is None:
        return

    # Buscar el nombre en la primera columna y actualizar la marca 'si' en la columna de la fecha actual
    encontrado = False
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == nombre:
                encontrado = True
                # Obtener la fila del nombre
                fila = cell.row

                # Actualizar la celda correspondiente a la fecha actual a 'si'
                ws.cell(row=fila, column=columna_fecha).value = 'si'
                break
    
    # Si el nombre no fue encontrado, agregar un nuevo alumno con la marca 'si' en la columna de la fecha actual
    if not encontrado:
        Nuevo_alumno(ws, nombre, columna_fecha)


def Nuevo_alumno(ws, nombre, columna_fecha):
    # Buscar la primera fila vacía en la columna de nombres
    fila_vacia = 2  # Empezamos en la fila 2, donde comienza la data
    while ws.cell(row=fila_vacia, column=1).value is not None:
        fila_vacia += 1
    ws.cell(row=fila_vacia, column=1).value = nombre

    # Actualizar la celda correspondiente a la fecha actual a 'si'
    ws.cell(row=fila_vacia, column=columna_fecha).value = 'si'


        

def marcar_asistencia_final():
    # Cargar el archivo Excel con openpyxl
    wb = load_workbook('Asistencia.xlsx')
    ws = wb.active
    
    # Obtener la fecha actual en el mismo formato que las fechas en el archivo Excel
    fecha_actual = datetime.now().strftime('%d/%m/%Y')

    columna_fecha = None

    # Buscar la columna correspondiente a la fecha actual
    for col in range(2, ws.max_column + 1):
        # Obtener el valor de la celda
        fecha_celda = ws.cell(row=1, column=col).value
        # Verificar si el valor de la celda es una cadena de texto
        if isinstance(fecha_celda, str):
            # Comparar la fecha de la hoja con la fecha actual
            if fecha_celda == fecha_actual:
                columna_fecha = col
                break
        elif isinstance(fecha_celda, datetime):
            # Convertir la fecha de la celda a cadena de texto y comparar con la fecha actual
            fecha_hoja = fecha_celda.strftime('%d/%m/%Y')
            if fecha_hoja == fecha_actual:
                columna_fecha = col
                break

    if columna_fecha is None:
        # Si la fecha actual no está en la hoja de cálculo, salir
        return

    # Iterar sobre todas las filas desde la segunda fila hasta la última
    for fila in range(2, ws.max_row + 1):
        # Verificar si la columna principal (columna 1) contiene un nombre
        if ws.cell(row=fila, column=1).value is not None:
            if ws.cell(row=fila, column=columna_fecha).value is None:
             # Escribir "no" en la celda correspondiente a la fecha actual
                ws.cell(row=fila, column=columna_fecha).value = 'no'

    # Guardar los cambios en el archivo Excel
    wb.save('Asistencia.xlsx')


       

def asistencia(nombre):
    # Cargar el archivo Excel con openpyxl
    wb = load_workbook('Asistencia.xlsx')
    ws = wb.active

    marca_asistencia(ws, nombre)
    

    # Guardar los cambios en el archivo Excel
    wb.save('Asistencia.xlsx')
   




def completar_fechas_vacias():
    # Cargar el archivo Excel con openpyxl
    wb = load_workbook('Asistencia.xlsx')
    ws = wb.active
    
    # Obtener la fecha actual en el mismo formato que las fechas en el archivo Excel
    fecha_actual = datetime.now().strftime('%d/%m/%Y')

    columna_fecha = None

    # Buscar la columna correspondiente a la fecha actual
    for col in range(2, ws.max_column + 1):
        # Obtener el valor de la celda
        fecha_celda = ws.cell(row=1, column=col).value
        # Verificar si el valor de la celda es una cadena de texto
        if isinstance(fecha_celda, str):
            # Comparar la fecha de la hoja con la fecha actual
            if fecha_celda == fecha_actual:
                columna_fecha = col
                break

    if columna_fecha is None:
        # Si la fecha actual no está en la hoja de cálculo, salir
        return

    # Iterar sobre todas las filas desde la segunda fila hasta la última
    for fila in range(2, ws.max_row + 1):
        # Verificar si la columna 0 contiene un valor (es decir, no es None)
        if ws.cell(row=fila, column=1).value is not None:
            # Iterar desde la columna de la fecha hasta la columna 2 (exclusiva)
            for col in range(columna_fecha, 1, -1):
                # Verificar si la celda está vacía
                if ws.cell(row=fila, column=col).value is None:
                    # Escribir "no" en la celda vacía
                    ws.cell(row=fila, column=col).value = 'no'
                
    # Guardar los cambios en el archivo Excel
    wb.save('Asistencia.xlsx')



      